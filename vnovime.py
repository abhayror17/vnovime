import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, NamedStyle
from datetime import timedelta
import glob
import xlsxwriter

# --- New ASCII Art Banner ---
ascii_art = r"""
╔══════════════════════════════════════════════════════════════════╗
║                                                                  ║
║   ██╗   ██╗███╗   ██╗ ██████╗ ██╗   ██╗██╗███╗   ███╗███████╗    ║
║   ██║   ██║████╗  ██║██╔═══██╗██║   ██║██║████╗ ████║██╔════╝    ║
║   ██║   ██║██╔██╗ ██║██║   ██║██║   ██║██║██╔████╔██║█████╗      ║
║   ██║   ██║██║╚██╗██║██║   ██║╚██╗ ██╔╝██║██║╚██╔╝██║██╔══╝      ║
║   ╚██████╔╝██║ ╚████║╚██████╔╝ ╚████╔╝ ██║██║ ╚═╝ ██║███████╗    ║
║    ╚═════╝ ╚═╝  ╚═══╝ ╚═════╝   ╚═══╝  ╚═╝╚═╝     ╚═╝╚══════╝    ║
║                                                                  ║
║                            V N O V I M E                         ║
║                                                                  ║
╚══════════════════════════════════════════════════════════════════╝ 
"""
print(ascii_art)
# --- End of Banner ---

# --- Configuration ---
# Automatically detect Excel files in the current directory
excel_files = glob.glob(os.path.join(os.getcwd(), '*.xlsx'))
excel_files = [f for f in excel_files if os.path.basename(f) != 'Aligned_Report_Sorted.xlsx']  # Exclude output file

# Identify the big and small files based on file size
if len(excel_files) >= 2:
    # Sort by file size (descending) to get the biggest file first
    excel_files.sort(key=lambda x: os.path.getsize(x), reverse=True)
    big_file_path = excel_files[0]
    small_file_path = excel_files[1]
    print(f"Big file: {os.path.basename(big_file_path)} ({os.path.getsize(big_file_path) / (1024*1024):.2f} MB)")
    print(f"Small file: {os.path.basename(small_file_path)} ({os.path.getsize(small_file_path) / (1024*1024):.2f} MB)")
else:
    print("Error: Need at least 2 Excel files in the directory")
    exit(1)

sheet_name = 'expressreport'
# Create output filename based on big file name, removing any (1), (2), etc. and adding "Line Items Added"
big_file_basename = os.path.splitext(os.path.basename(big_file_path))[0]
# Remove any (1), (2), etc. from the end of the filename
import re
big_file_basename = re.sub(r'\s*\(\d+\)$', '', big_file_basename)
output_file = f'{big_file_basename} Line Items Added.xlsx'

# --- 1. Load the Data ---
print("Loading files...")
try:
    df_big = pd.read_excel(big_file_path, sheet_name=sheet_name)
    print(f"  Successfully read {len(df_big)} rows from big file.")
except ValueError:
    print(f"  Error: Worksheet '{sheet_name}' not found in {os.path.basename(big_file_path)}")
    exit(1)
except Exception as e:
    print(f"  Error reading {os.path.basename(big_file_path)}: {str(e)}")
    exit(1)

try:
    df_small = pd.read_excel(small_file_path, sheet_name=sheet_name)
    print(f"  Successfully read {len(df_small)} rows from small file.")
except ValueError:
    print(f"  Error: Worksheet '{sheet_name}' not found in {os.path.basename(small_file_path)}")
    exit(1)
except Exception as e:
    print(f"  Error reading {os.path.basename(small_file_path)}: {str(e)}")
    exit(1)

# --- 2. Identify Missing Rows ---
df_big['Sr No'] = df_big['Sr No'].astype(str)
df_small['Sr No'] = df_small['Sr No'].astype(str)

existing_ids = set(df_big['Sr No'])
new_rows = df_small[~df_small['Sr No'].isin(existing_ids)].copy()

print(f"Found {len(new_rows)} new rows to add.")

if not new_rows.empty:
    # --- 3. Align Columns ---
    # We want the new rows to look EXACTLY like the big file's rows.
    # So we force 'new_rows' to have the same columns as 'df_big'.
    
    big_columns = df_big.columns.tolist()
    
    for col in big_columns:
        if col not in new_rows.columns:
            # If the column (e.g., 'Story Ori') is missing in the small file,
            # create it in the new rows and leave it empty (None).
            new_rows[col] = None
            print(f"--> Created empty column '{col}' in new rows to match structure.")
            
    # Reorder the columns of new_rows to match df_big exactly
    # This ensures no "misplacement" happens.
    # Note: This also drops any extra columns from the small file that aren't in the big file.
    new_rows = new_rows[big_columns]

    # --- 4. Append and Sort ---
    print("Starting merge process...")
    
    # Mark new rows BEFORE concatenation
    df_big["__is_new"] = False
    new_rows["__is_new"] = True
    
    # Combine
    df_combined = pd.concat([df_big, new_rows], ignore_index=True)
    print(f"Successfully merged {len(df_combined)} total rows.")
    
    # Sort by channel name, program date, and clip start time if these columns exist
    sort_columns = []
    if 'Channel Name' in df_combined.columns:
        sort_columns.append('Channel Name')
    if 'Program Date' in df_combined.columns:
        sort_columns.append('Program Date')
    if 'Clip Start Time' in df_combined.columns:
        sort_columns.append('Clip Start Time')
        
    # Sort (if applicable)
    if sort_columns:
        print(f"Sorting by columns: {sort_columns}")
        df_combined = df_combined.sort_values(by=sort_columns)
    
    # Reset index so positions match Excel rows
    df_combined = df_combined.reset_index(drop=True)
    
    # Identify positions of new rows AFTER sorting
    new_row_positions = df_combined.index[df_combined["__is_new"] == True].tolist()
    
    # Remove temp flag
    df_combined = df_combined.drop(columns=["__is_new"])
    
    # --- 5. Save with Formatting ---
    try:
        print(f"Attempting to save and format the file as '{os.path.basename(output_file)}'...")
        with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
            df_combined.to_excel(writer, sheet_name=sheet_name, index=False, header=False, startrow=1)
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Updated header format to include center alignment
            header_format = workbook.add_format({
                'bold': True,
                'align': 'center',      # Center horizontally
                'valign': 'vcenter',    # Center vertically
                'fg_color': '#6F6F6F',
                'font_color': 'white',
                'border': 1
            })
            
            for col_num, value in enumerate(df_combined.columns.values):
                worksheet.write(0, col_num, value, header_format)
                column_len = max(
                    len(str(value)),
                    df_combined[value].astype(str).str.len().max()
                )
                worksheet.set_column(col_num, col_num, column_len + 4)  # Added a bit more padding for centered text
                
        # --- 6. Highlight New Rows ---
        print("Highlighting new rows...")
        wb = load_workbook(output_file)
        ws = wb[sheet_name]
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
        # Highlight the new rows (accounting for header row and 0-based indexing)
        for idx in new_row_positions:
            excel_row = idx + 2  # +1 for header, +1 for 1-based Excel indexing
            for cell in ws[excel_row]:
                cell.fill = yellow_fill
        
        # --- 7. Apply Duration Formatting (from pivot.py) ---
        print("Applying duration formatting...")
        
        # Define the time style for Excel (hh:mm:ss) once
        time_style = NamedStyle(name="time_style", number_format="hh:mm:ss")
        
        # Find the Duration column index
        duration_col_idx = None
        for col_idx, column_name in enumerate(df_combined.columns):
            if 'Duration' in str(column_name):
                duration_col_idx = col_idx + 1  # +1 for 1-based Excel indexing
                print(f"Found Duration column at index {duration_col_idx}: {column_name}")
                break
        
        if duration_col_idx:
            # Apply the time style to the Duration column
            for row in ws.iter_rows(min_row=2, min_col=duration_col_idx, max_col=duration_col_idx):
                for cell in row:
                    # Check if the cell value is a string that looks like a duration
                    if isinstance(cell.value, str) and ':' in str(cell.value):
                        try:
                            # Convert duration string to timedelta for proper Excel formatting
                            h, m, s = map(int, str(cell.value).split(':'))
                            cell.value = timedelta(hours=h, minutes=m, seconds=s)
                            cell.style = time_style
                        except:
                            pass  # Skip if conversion fails
                    elif isinstance(cell.value, (int, float)):
                        # If it's already a number, assume it's seconds and convert
                        try:
                            seconds = int(cell.value)
                            hours = seconds // 3600
                            minutes = (seconds % 3600) // 60
                            secs = seconds % 60
                            cell.value = timedelta(hours=hours, minutes=minutes, seconds=secs)
                            cell.style = time_style
                        except:
                            pass
                
        wb.save(output_file)
                    
        print("\nChecking if file was created...")
        if os.path.exists(output_file):
            size_mb = os.path.getsize(output_file) / (1024 * 1024)
            print(f"Success! Output file created: {output_file}")
            print(f"File size: {size_mb:.2f} MB")
        else:
            print("Error: File was not created")
            
    except PermissionError:
        print(f"\nError: Permission denied. Please make sure '{os.path.basename(output_file)}' is not open.")
    except MemoryError:
        print("\nError: Not enough memory to complete the operation.")
    except Exception as e:
        print(f"\nAn error occurred while writing the merged file: {str(e)}")
        
else:
    print("No missing rows found.")

